from openpyxl import Workbook, load_workbook
from googlesearch import GoogleSearch
wb = load_workbook('Attendees.xlsx')
for sheet in wb:
	ws = wb[sheet.title]
	for i in range(1, 290):
		nameCell = ws.cell(row=i, column=1)
		jobCell = ws.cell(row=i, column=3)
		search = str(nameCell.value) + " " + str(jobCell.value) + " email"
		gs = GoogleSearch(search)
		print search
		print gs.top_urls()[0]
		print("\n")